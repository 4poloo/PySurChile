import os
import json
import pandas as pd
import openpyxl
from datetime import datetime

CONFIG_FILE = "config.json"

# Función para leer o crear el archivo de configuración
def obtener_config():
    if not os.path.exists(CONFIG_FILE):
        # Si no existe el archivo de configuración, lo creamos con valores iniciales
        with open(CONFIG_FILE, "w") as f:
            json.dump({"ultimo_folio": 1000, "ultimo_registro": "", "ultima_hora": "00:00:00"}, f)
    with open(CONFIG_FILE, "r") as f:
        return json.load(f)

# Función para actualizar la configuración (fecha y hora)
def actualizar_config(registro, hora):
    config = obtener_config()
    config["ultimo_registro"] = registro
    config["ultima_hora"] = hora
    with open(CONFIG_FILE, "w") as f:
        json.dump(config, f, indent=4)

# Función para actualizar solo el folio
def actualizar_folio(folio):
    config = obtener_config()
    config["ultimo_folio"] = folio
    with open(CONFIG_FILE, "w") as f:
        json.dump(config, f, indent=4)

# Función para leer el archivo Excel
def leer_excel(ruta_archivo):
    try:
        df = pd.read_excel(ruta_archivo, dtype=str)
        return df
    except Exception as e:
        raise ValueError(f"Error al leer el archivo Excel: {str(e)}")

# Función para validar las columnas requeridas
def validar_columnas(df):
    columnas_requeridas = ['PRODUCTO', 'OT', 'UNID. DISP.', 'FECHA RECIBO']
    if not all(col in df.columns for col in columnas_requeridas):
        raise ValueError(f"El archivo debe contener las columnas: {', '.join(columnas_requeridas)}")
    return df[columnas_requeridas]

# Función para procesar los datos, ahora con manejo de la hora
def procesar_datos(df, ultimo_registro, ultima_hora):
    # Convertir la columna de fecha (que tiene fecha y hora) en tipo datetime
    df['FECHA RECIBO'] = pd.to_datetime(df['FECHA RECIBO'], format='%d/%m/%Y %H:%M:%S', errors='coerce')

    # Filtrar solo los registros que sean posteriores a la última fecha y hora
    if ultimo_registro and ultima_hora:
        fecha_registro = pd.to_datetime(ultimo_registro, dayfirst=True)
        hora_registro = datetime.strptime(ultima_hora, "%H:%M:%S").time()

        # Combinar fecha y hora para una comparación precisa
        fecha_y_hora_registro = datetime.combine(fecha_registro, hora_registro)

        # Filtrar los registros que son posteriores al último registro (fecha + hora)
        df['FECHA_RECIBO_HORA'] = pd.to_datetime(df['FECHA RECIBO'].astype(str))
        df = df[df['FECHA_RECIBO_HORA'] > fecha_y_hora_registro]

    # Filtrar y convertir datos a numéricos
    df['UNID. DISP.'] = pd.to_numeric(df['UNID. DISP.'], errors='coerce')
    df['OT'] = df['OT'].apply(lambda x: x[3:] if isinstance(x, str) and x.startswith('OT-') else x)

    return df

# Función para agrupar los datos
def agrupar_datos(df):
    # Agrupar los datos por 'PRODUCTO' y 'OT'
    df_agrupado = df.groupby(['PRODUCTO', 'OT']).agg({
        'UNID. DISP.': 'sum',
        'FECHA RECIBO': 'min'
    }).reset_index()
    return df_agrupado

# Generar el archivo Excel de salida
def generar_excel(df_agrupado, ruta_plantilla, ruta_salida, fecha_primera, hora_primera):
    try:
        wb = openpyxl.load_workbook(ruta_plantilla)
        ws = wb.active
    except Exception as e:
        raise ValueError(f"Error al cargar la plantilla: {str(e)}")

    # Limpiar filas existentes
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=23):
        for cell in row:
            cell.value = None

    # Obtener el último folio de la configuración
    config = obtener_config()
    ultimo_folio = config["ultimo_folio"]
    
    fecha_primera = pd.to_datetime(fecha_primera).strftime('%d/%m/%Y')
    fecha_guarda = pd.to_datetime(fecha_primera).strftime('%d_%m')
    hora_primera = datetime.strptime(hora_primera, "%H:%M:%S").strftime("%H:%M:%S")

    # Agregar datos al archivo de plantilla
    for i, (producto, ot, cantidad, fecha) in enumerate(df_agrupado.values, start=1):
        fila = 2 + i - 1
        ws[f"A{fila}"] = i
        ws[f"B{fila}"] = ultimo_folio + i - 1
        ws[f"U{fila}"] = producto
        ws[f"W{fila}"] = cantidad
        ws[f"O{fila}"] = ot
        ws[f"D{fila}"] = 10
        ws[f"C{fila}"] = fecha.strftime('%d/%m/%Y')
        ws[f"F{fila}"] = 77757460

    # Guardar archivo
    nuevo_folio = ultimo_folio + len(df_agrupado)

    # Actualizar la configuración con el nuevo folio antes de guardar el archivo
    actualizar_folio(nuevo_folio)
    
    nombre_salida = f"{ruta_salida}_{fecha_guarda}.xlsx"
    wb.save(nombre_salida)

    return nombre_salida

# Función principal que orquesta todo el proceso
def procesar_archivo(ruta_entrada, ruta_plantilla, ruta_salida):
    try:
        config = obtener_config()
        ultimo_registro = config.get("ultimo_registro", "")
        ultima_hora = config.get("ultima_hora", "")

        # Leer y validar el archivo Excel
        df = leer_excel(ruta_entrada)
        df = validar_columnas(df)

        # Tomar la primera fila de FECHA RECIBO antes de procesar los datos
        fecha_primera_sin_procesar = pd.to_datetime(df['FECHA RECIBO']).max()
        hora_primera_sin_procesar = fecha_primera_sin_procesar.strftime('%H:%M:%S')
        fecha_primera_sin_procesar = fecha_primera_sin_procesar.strftime('%d/%m/%Y')

        # Procesar y filtrar datos
        df = procesar_datos(df, ultimo_registro, ultima_hora)
        if df.empty:
            return {"mensaje": "No se han procesado datos nuevos.", "archivo_salida": None}

        # Agrupar los datos
        df_agrupado = agrupar_datos(df)

        # Tomar la fecha y hora más reciente después de procesar los datos
        fecha_ultima = df_agrupado['FECHA RECIBO'].max()
        hora_ultima = fecha_ultima.strftime('%H:%M:%S')
        fecha_ultima = fecha_ultima.strftime('%d/%m/%Y')

        # Generar el archivo Excel de salida
        archivo_salida = generar_excel(df_agrupado, ruta_plantilla, ruta_salida, fecha_ultima, hora_ultima)

        # Actualizar config con la última fecha y hora del archivo original
        actualizar_config(fecha_primera_sin_procesar, hora_primera_sin_procesar)

        return {"mensaje": "Procesamiento exitoso", "archivo_salida": archivo_salida}

    except Exception as e:
        return {"error": str(e)}

# Bloque principal para ejecución
if __name__ == "__main__":
    # Definir las rutas de los archivos
    ruta_entrada = "./HistorialCaja.xlsx"  # Aquí pon la ruta de tu archivo de entrada
    print("Ruta al archivo:", os.path.abspath(ruta_entrada))
    ruta_plantilla = "./Plantilla/Plantilla_ERP.xlsx"  # Ruta de tu plantilla
    ruta_salida = "./Plantilla/CARGA_PT"  # Ruta donde se guardará el archivo de salida
    
    # Llamar la función para procesar el archivo
    resultado = procesar_archivo(ruta_entrada, ruta_plantilla, ruta_salida)

    # Imprimir el resultado
    if "error" in resultado:
        print(f"Error: {resultado['error']}")
    else:
        print(f"Éxito: {resultado['mensaje']}")
        if resultado['archivo_salida']:
            print(f"Archivo generado: {resultado['archivo_salida']}")
        else:
            print("No se generó un archivo, ya que no se procesaron datos nuevos.")
