from fastapi import FastAPI, UploadFile, File
from pydantic import BaseModel
import os
import json
import pandas as pd
import openpyxl
from datetime import datetime
from io import BytesIO

app = FastAPI()

CONFIG_FILE = "config.json"
TEMPLATE_FILE = "./Plantilla/Plantilla_ERP.xlsx"

# Función para leer o crear el archivo de configuración
def obtener_config():
    if not os.path.exists(CONFIG_FILE):
        # Si no existe el archivo de configuración, lo creamos con valores iniciales
        with open(CONFIG_FILE, "w") as f:
            json.dump({"ultimo_folio": 1000, "ultimo_registro": "", "ultima_hora": "00:00:00"}, f)
    with open(CONFIG_FILE, "r") as f:
        return json.load(f)

# Función para actualizar la configuración (fecha y hora)
def actualizar_config(registro, hora):
    config = obtener_config()
    config["ultimo_registro"] = registro
    config["ultima_hora"] = hora
    with open(CONFIG_FILE, "w") as f:
        json.dump(config, f, indent=4)

# Función para actualizar solo el folio
def actualizar_folio(folio):
    config = obtener_config()
    config["ultimo_folio"] = folio
    with open(CONFIG_FILE, "w") as f:
        json.dump(config, f, indent=4)

# Función para leer el archivo Excel
def leer_excel(ruta_archivo):
    try:
        df = pd.read_excel(ruta_archivo, dtype=str)
        return df
    except Exception as e:
        raise ValueError(f"Error al leer el archivo Excel: {str(e)}")

# Función para validar las columnas requeridas
def validar_columnas(df):
    columnas_requeridas = ['PRODUCTO', 'OT', 'UNID. DISP.', 'FECHA RECIBO']
    if not all(col in df.columns for col in columnas_requeridas):
        raise ValueError(f"El archivo debe contener las columnas: {', '.join(columnas_requeridas)}")
    return df[columnas_requeridas]

# Función para procesar los datos, ahora con manejo de la hora
def procesar_datos(df, ultimo_registro, ultima_hora):
    df['FECHA RECIBO'] = pd.to_datetime(df['FECHA RECIBO'], format='%d/%m/%Y %H:%M:%S', errors='coerce')

    if ultimo_registro and ultima_hora:
        fecha_registro = pd.to_datetime(ultimo_registro, dayfirst=True)
        hora_registro = datetime.strptime(ultima_hora, "%H:%M:%S").time()

        fecha_y_hora_registro = datetime.combine(fecha_registro, hora_registro)

        df['FECHA_RECIBO_HORA'] = pd.to_datetime(df['FECHA RECIBO'].astype(str))
        df = df[df['FECHA_RECIBO_HORA'] > fecha_y_hora_registro]

    df['UNID. DISP.'] = pd.to_numeric(df['UNID. DISP.'], errors='coerce')
    df['OT'] = df['OT'].apply(lambda x: x[3:] if isinstance(x, str) and x.startswith('OT-') else x)

    return df

# Función para agrupar los datos
def agrupar_datos(df):
    df_agrupado = df.groupby(['PRODUCTO', 'OT']).agg({
        'UNID. DISP.': 'sum',
        'FECHA RECIBO': 'min'
    }).reset_index()
    return df_agrupado

# Generar el archivo Excel de salida
def generar_excel(df_agrupado, ruta_plantilla, ruta_salida, fecha_primera, hora_primera):
    try:
        wb = openpyxl.load_workbook(ruta_plantilla)
        ws = wb.active
    except Exception as e:
        raise ValueError(f"Error al cargar la plantilla: {str(e)}")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=23):
        for cell in row:
            cell.value = None

    config = obtener_config()
    ultimo_folio = config["ultimo_folio"]
    
    fecha_primera = pd.to_datetime(fecha_primera).strftime('%d/%m/%Y')
    fecha_guarda = pd.to_datetime(fecha_primera).strftime('%d_%m')
    hora_primera = datetime.strptime(hora_primera, "%H:%M:%S").strftime("%H:%M:%S")

    for i, (producto, ot, cantidad, fecha) in enumerate(df_agrupado.values, start=1):
        fila = 2 + i - 1
        ws[f"A{fila}"] = i
        ws[f"B{fila}"] = ultimo_folio + i - 1
        ws[f"U{fila}"] = producto
        ws[f"W{fila}"] = cantidad
        ws[f"O{fila}"] = ot
        ws[f"D{fila}"] = 10
        ws[f"C{fila}"] = fecha.strftime('%d/%m/%Y')
        ws[f"F{fila}"] = 77757460

    nuevo_folio = ultimo_folio + len(df_agrupado)
    actualizar_folio(nuevo_folio)
    
    nombre_salida = f"{ruta_salida}_{fecha_guarda}.xlsx"
    wb.save(nombre_salida)

    return nombre_salida

# Función principal para procesar el archivo Excel
def procesar_archivo(df):
    try:
        config = obtener_config()
        ultimo_registro = config.get("ultimo_registro", "")
        ultima_hora = config.get("ultima_hora", "")

        df = validar_columnas(df)
        fecha_primera_sin_procesar = pd.to_datetime(df['FECHA RECIBO']).max()
        hora_primera_sin_procesar = fecha_primera_sin_procesar.strftime('%H:%M:%S')
        fecha_primera_sin_procesar = fecha_primera_sin_procesar.strftime('%d/%m/%Y')

        df = procesar_datos(df, ultimo_registro, ultima_hora)
        if df.empty:
            return {"mensaje": "No se han procesado datos nuevos.", "archivo_salida": None}

        df_agrupado = agrupar_datos(df)

        fecha_ultima = df_agrupado['FECHA RECIBO'].max()
        hora_ultima = fecha_ultima.strftime('%H:%M:%S')
        fecha_ultima = fecha_ultima.strftime('%d/%m/%Y')

        archivo_salida = generar_excel(df_agrupado, TEMPLATE_FILE, "./Plantilla/CARGA_PT", fecha_ultima, hora_ultima)

        actualizar_config(fecha_primera_sin_procesar, hora_primera_sin_procesar)

        return {"mensaje": "Procesamiento exitoso", "archivo_salida": archivo_salida}

    except Exception as e:
        return {"error": str(e)}

