import pandas as pd
import openpyxl
import json
import os

# Ruta del archivo de configuración
CONFIG_FILE = "config.json"

# Función para leer o inicializar el archivo de configuración
def obtener_ultimo_folio():
    if not os.path.exists(CONFIG_FILE):
        # Crear archivo de configuración si no existe
        with open(CONFIG_FILE, "w") as config:
            json.dump({"ultimo_folio": 1000, "ultimo_registro": ""}, config)
        return 1000
    else:
        with open(CONFIG_FILE, "r") as config:
            data = json.load(config)
        return data["ultimo_folio"]

# Función para obtener la última fecha registrada (corte) del JSON
def obtener_ultimo_registro():
    if not os.path.exists(CONFIG_FILE):
        return ""
    else:
        with open(CONFIG_FILE, "r") as config:
            data = json.load(config)
        return data["ultimo_registro"]

# Función para actualizar el archivo de configuración con el nuevo folio y fecha
def actualizar_ultimo_folio(nuevo_folio, ultimo_registro):
    # Asegurarse de que el último registro sea una cadena (por si es un objeto Timestamp)
    if not isinstance(ultimo_registro, str):
        ultimo_registro = ultimo_registro.strftime('%Y-%m-%d %H:%M:%S')
    with open(CONFIG_FILE, "w") as config:
        json.dump({"ultimo_folio": nuevo_folio, "ultimo_registro": ultimo_registro}, config)

# Función para procesar el archivo Excel de entrada
def procesar_archivo_excel(ruta_archivo, ultimo_registro):
    try:
        # Leer el archivo Excel
        df = pd.read_excel(ruta_archivo, engine='openpyxl', dtype=str)
    except Exception as e:
        print(f"Error al leer el archivo: {e}")
        return None, None

    columnas_esperadas = ['PRODUCTO', 'OT', 'UNID. DISP.', 'FECHA RECIBO']
    if not all(col in df.columns for col in columnas_esperadas):
        print(f"El archivo debe contener las columnas: {', '.join(columnas_esperadas)}")
        return None, None

    # Convertir 'UNID. DISP.' a numérico para asegurar que las sumas sean correctas
    df['UNID. DISP.'] = pd.to_numeric(df['UNID. DISP.'], errors='coerce')

    # Eliminar el prefijo 'OT-' si existe
    df['OT'] = df['OT'].apply(lambda x: x[3:] if isinstance(x, str) and x.startswith('OT-') else x)

    # Convertir 'FECHA RECIBO' a tipo datetime
    df['FECHA RECIBO'] = pd.to_datetime(df['FECHA RECIBO'], errors='coerce')

    # Filtrar filas con fecha posterior al último registro
    if ultimo_registro:
        fecha_registrada = pd.to_datetime(ultimo_registro)
        df = df[df['FECHA RECIBO'] > fecha_registrada]

    # Si el dataframe está vacío después del filtro, no continuar
    if df.empty:
        print("No hay registros nuevos para procesar.")
        return None, None

    # Ordenar por fecha descendente
    df = df.sort_values(by='FECHA RECIBO', ascending=False)

    # Obtener la fecha y hora de la primera fila como referencia
    fecha_referencia = df.iloc[0]['FECHA RECIBO']
    print(f"Fecha de referencia seleccionada: {fecha_referencia}")

    # Agrupar por 'PRODUCTO' y 'OT' y sumar 'UNID. DISP.'
    resultado_agrupado = df.groupby(['PRODUCTO', 'OT'], as_index=False).agg({'UNID. DISP.': 'sum', 'FECHA RECIBO': 'min'})

    # Convertir a una lista de listas (array) con los valores procesados
    resultado = resultado_agrupado.values.tolist()

    # Devolver los datos procesados y la fecha de referencia
    return resultado, fecha_referencia.strftime('%Y-%m-%d %H:%M:%S')

# Función para obtener la fecha completa (fecha + hora) del último registro en el archivo Excel
def obtener_fecha_completa_ultimo_registro(ruta_archivo):
    try:
        # Leer el archivo Excel
        df = pd.read_excel(ruta_archivo, engine='openpyxl', dtype=str)

        # Asegurarnos de que la columna 'FECHA RECIBO' exista
        if 'FECHA RECIBO' not in df.columns:
            print("El archivo no contiene la columna 'FECHA RECIBO'.")
            return None
        
        # Convertir la columna 'FECHA RECIBO' a tipo datetime
        df['FECHA RECIBO'] = pd.to_datetime(df['FECHA RECIBO'], errors='coerce')
        
        # Ordenar por fecha descendente para obtener la primera fila
        df = df.sort_values(by='FECHA RECIBO', ascending=False)
        
        # Obtener la fecha completa (fecha + hora) de la primera fila
        if not df.empty:
            fecha_ultimo_registro = df.iloc[0]['FECHA RECIBO']
            # Devolver la fecha y hora en formato completo (YYYY-MM-DD HH:MM:SS)
            return fecha_ultimo_registro.strftime('%Y-%m-%d %H:%M:%S')
        else:
            print("No hay registros en el archivo.")
            return None
    except Exception as e:
        print(f"Error al procesar el archivo: {e}")
        return None

# Rellenar la plantilla con los datos procesados
def rellenar_plantilla(datos_array, ruta_plantilla, ruta_salida_base):
    # Intentar cargar la plantilla de Excel
    try:
        wb = openpyxl.load_workbook(ruta_plantilla)
        ws = wb.active  # Usar la primera hoja activa
    except Exception as e:
        print(f"Error al cargar la plantilla: {e}")
        return

    # Limpiar los datos existentes en la plantilla
    fila_inicio = 2  # Supongamos que los datos empiezan en la fila 2
    for row in ws.iter_rows(min_row=fila_inicio, max_row=ws.max_row, min_col=2, max_col=23):
        for cell in row:
            cell.value = None

    # Obtener el último folio guardado en el archivo de configuración
    ultimo_folio = obtener_ultimo_folio()

    # Rellenar la plantilla con los datos nuevos
    for i, (producto, ot, cantidad, fecha) in enumerate(datos_array, start=1):
        fila_actual = fila_inicio + i - 1  # Calcular la fila actual en la hoja
        fecha_formateada = pd.to_datetime(fecha).strftime('%d/%m/%Y')  # Formato dd/mm/yyyy
        ws[f"B{fila_actual}"] = ultimo_folio + i - 1  # Número de folio
        ws[f"U{fila_actual}"] = producto  # Código de Producto
        ws[f"W{fila_actual}"] = cantidad  # Cantidad Ingresada
        ws[f"O{fila_actual}"] = ot  # Número de Orden de Producción (Interna)
        ws[f"D{fila_actual}"] = 10  # Concepto de Entrada a Bodega
        ws[f"C{fila_actual}"] = fecha_formateada  # Fecha de Generación de Guía de Entrada
        ws[f"F{fila_actual}"] = 77757460  # Código de Proveedor

    # Guardar el archivo de salida con el formato deseado
    try:
        nuevo_folio = ultimo_folio + len(datos_array)  # Calcular el nuevo folio
        if datos_array:
            # Guardar la fecha del registro más reciente en el archivo de configuración
            ultimo_registro = obtener_fecha_completa_ultimo_registro(ruta_entrada)  # Se usa la fecha del primer registro como último registro
            actualizar_ultimo_folio(nuevo_folio, ultimo_registro)

            # Crear el nombre del archivo de salida con formato CARGA_PT_DD_MM.xlsx
            fecha_registro = pd.to_datetime(ultimo_registro)
            nombre_salida = f"{ruta_salida_base}_{fecha_registro.strftime('%d_%m')}.xlsx"
            wb.save(nombre_salida)
            print(f"Archivo exportado exitosamente: {nombre_salida}")
    except Exception as e:
        print(f"Error al guardar el archivo exportado: {e}")

if __name__ == "__main__":
    ruta_entrada = input("Ingresa la ruta al archivo Excel de entrada: ")
    ultimo_registro = obtener_ultimo_registro()
    datos, nuevo_registro = procesar_archivo_excel(ruta_entrada, ultimo_registro)
    if datos:
        ruta_plantilla = "./plantilla/Plantilla_ERP.xlsx"
        ruta_salida_base = "./plantilla/CARGA_PT"
        rellenar_plantilla(datos, ruta_plantilla, ruta_salida_base)
        print(f"Último registro procesado: {nuevo_registro}")
    else:
        print("No se han procesado datos nuevos.")
